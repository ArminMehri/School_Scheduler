from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from main.services.validator import run_full_validation
from main.models import *
from .solver_engine import generate_schedule_with_ortools

from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml import OxmlElement

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

from main.services.pdf_export import *
import arabic_reshaper
from bidi.algorithm import get_display

from django.conf import settings
from django.utils import timezone
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.contrib.staticfiles import finders
import os


def fa(text: str) -> str:
    if text is None:
        text = ""
    reshaped = arabic_reshaper.reshape(str(text))
    return get_display(reshaped)


def _require_school(request):
    school = getattr(request.user, "school", None)
    if school is None:
        messages.error(request, "❌ این حساب به هیچ مدرسه‌ای وصل نیست.")
    return school

def company_index(request):
    index_title1 = announce_company.objects.filter(announce=True).order_by('-id').first()

    context = {
        'index_title1': index_title1,
    }
    return render(request, "site/company_index.html", context)

def about(request):
    return render(request, "site/about.html")
User = get_user_model()

# =========================
# Account Management Page
# =========================
@staff_member_required
def account_manage(request):

    users = User.objects.all().order_by('-id')

    total_users = users.count()

    active_users = users.filter(
        is_active=True
    ).count()

    blocked_users = users.filter(
        is_active=False
    ).count()

    staff_users = users.filter(
        is_staff=True
    ).count()

    context = {
        'users': users,

        'total_users': total_users,
        'active_users': active_users,
        'blocked_users': blocked_users,
        'staff_users': staff_users,
    }

    return render(
        request,
        'site/account_manage.html',
        context
    )


# =========================
# Cart Page
# =========================
def cart_page(request):

    """
    نمونه سبد خرید موقت
    بعداً میتونی از session یا database بخونی
    """

    cart_items = [

        {
            'title': 'پنل برنامه‌ساز مدارس',
            'description': 'نسخه حرفه‌ای مدیریت مدارس',
            'quantity': 1,
            'price': 2500000,
            'total_price': 2500000,
        },

        {
            'title': 'طراحی سایت اختصاصی',
            'description': 'طراحی وب‌سایت شرکتی و مدیریتی',
            'quantity': 1,
            'price': 4800000,
            'total_price': 4800000,
        },

    ]

    subtotal = sum(
        item['total_price']
        for item in cart_items
    )

    discount = 500000

    tax = 0

    total = subtotal - discount + tax

    context = {
        'cart_items': cart_items,

        'subtotal': subtotal,
        'discount': discount,
        'tax': tax,
        'total': total,
    }

    return render(
        request,
        'site/cart.html',
        context
    )

def site_index(request):
    index_title = announce.objects.all().first()
    context = {
        'index_title': index_title,
    }

    return render(request, "index.html",context)


def export_schedule_excel(request):
    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "برنامه کلاسی"
    ws.sheet_view.rightToLeft = True

    # ---------- Styles ----------
    b_titr = Font(name="B Titr")

    align_right_center = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rotate Up (90 درجه)
    rotate_up = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=False)
    rotate_up_wrap = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

    # قرمز برای بدون دبیر
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold = Font(name="B Titr", color="FFFFFF", bold=True)

    # ---------- Data ----------
    classes = SchoolClass.objects.filter(school=school).select_related("grade").all()
    days = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods").all()

    # ---------- Header Row 1: Days ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1, value="نام کلاس")
    cell.alignment = align_right_center
    cell.font = b_titr

    col_index = 2
    for day in days:
        periods = list(day.periods.all())
        periods_count = len(periods)

        cell = ws.cell(row=1, column=col_index, value=day.name)
        cell.alignment = align_center  # روزها rotate نشوند (مثل قدیمی)
        cell.font = b_titr

        if periods_count > 1:
            ws.merge_cells(
                start_row=1, start_column=col_index,
                end_row=1, end_column=col_index + periods_count - 1
            )

        col_index += periods_count

    # ---------- Header Row 2: Periods (Rotate Up) ----------
    col_index = 2
    for day in days:
        for period in day.periods.all():
            cell = ws.cell(row=2, column=col_index, value=f"زنگ {period.period_number}")
            cell.font = b_titr
            cell.alignment = rotate_up  # ✅ rotate زنگ‌ها
            col_index += 1

    # کمی ارتفاع هدرها
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 55

    # ---------- محاسبه ستون شروع هر روز (برای خط جداکننده) ----------
    thick_side = Side(style="medium", color="000000")
    thin_side = Side(style="thin", color="CCCCCC")

    day_start_cols = []
    col_tracker = 2
    for day in days:
        day_start_cols.append(col_tracker)
        col_tracker += len(list(day.periods.all()))
    total_cols = col_tracker - 1

    # ---------- Body ----------
    row_index = 3
    for school_class in classes:
        cell_class = ws.cell(row=row_index, column=1, value=school_class.name)
        cell_class.alignment = align_right_center
        cell_class.font = b_titr

        col_index = 2
        for day in days:
            for period in day.periods.all():
                sched = Schedule.objects.filter(
                    school=school,
                    school_class=school_class,
                    day_period=period
                ).select_related("lesson", "teacher").first()

                cell = ws.cell(row=row_index, column=col_index)

                if sched:
                    cell.value = sched.lesson.name
                    if sched.teacher is None:
                        cell.fill = red_fill
                        cell.font = white_bold
                    else:
                        cell.font = b_titr
                    cell.alignment = rotate_up_wrap
                else:
                    cell.value = ""
                    cell.font = b_titr
                    cell.alignment = rotate_up

                left_b = thick_side if col_index in day_start_cols else thin_side
                right_b = thick_side if col_index == total_cols else thin_side
                cell.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thick_side)

                col_index += 1

        # خط پررنگ پایین ستون اول (نام کلاس) هم
        cell_class.border = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)

        ws.row_dimensions[row_index].height = 60
        row_index += 1

    # خط پررنگ روی هدرها
    for r in [1, 2]:
        col_index = 2
        for day in days:
            for period in day.periods.all():
                cell = ws.cell(row=r, column=col_index)
                left_b = thick_side if col_index in day_start_cols else thin_side
                right_b = thick_side if col_index == total_cols else thin_side
                cell.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thick_side)
                col_index += 1

    # ---------- Column widths ----------
    ws.column_dimensions["A"].width = 14
    max_col = ws.max_column
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5.2

    # ---------- Freeze panes ----------
    ws.freeze_panes = "B3"

    # ---------- Response ----------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=barname_kelasi.xlsx"
    wb.save(response)
    return response


def export_schedule_excel_teacher(request):
    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "برنامه دبیران"
    ws.sheet_view.rightToLeft = True

    # ---------- Styles ----------
    b_titr = Font(name="B Titr")

    align_right_center = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rotate_up = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=False)
    rotate_up_wrap = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

    # ---------- Data ----------
    teachers = Teacher.objects.filter(school=school).all()
    days = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods").all()

    # ---------- Header Row 1: Days ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1, value="نام دبیر")
    cell.alignment = align_right_center
    cell.font = b_titr

    col_index = 2
    for day in days:
        periods = list(day.periods.all())
        periods_count = len(periods)

        cell = ws.cell(row=1, column=col_index, value=day.name)
        cell.alignment = align_center  # روزها rotate نشوند
        cell.font = b_titr

        if periods_count > 1:
            ws.merge_cells(
                start_row=1, start_column=col_index,
                end_row=1, end_column=col_index + periods_count - 1
            )

        col_index += periods_count

    # ---------- Header Row 2: Periods (Rotate Up) ----------
    col_index = 2
    for day in days:
        for period in day.periods.all():
            cell = ws.cell(row=2, column=col_index, value=f"زنگ {period.period_number}")
            cell.font = b_titr
            cell.alignment = rotate_up
            col_index += 1

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 55

    # ---------- محاسبه ستون شروع هر روز (برای خط جداکننده) ----------
    thick_side = Side(style="medium", color="000000")
    thin_side = Side(style="thin", color="CCCCCC")

    day_start_cols = []
    col_tracker = 2
    for day in days:
        day_start_cols.append(col_tracker)
        col_tracker += len(list(day.periods.all()))
    total_cols = col_tracker - 1

    # ---------- Body ----------
    row_index = 3
    for teacher in teachers:
        cell_teacher = ws.cell(row=row_index, column=1, value=teacher.name)
        cell_teacher.alignment = align_right_center
        cell_teacher.font = b_titr

        col_index = 2
        for day in days:
            for period in day.periods.all():
                sched = Schedule.objects.filter(
                    school=school,
                    teacher=teacher,
                    day_period=period
                ).select_related("school_class", "lesson").first()

                cell = ws.cell(row=row_index, column=col_index)
                if sched:
                    cell.value = f"{sched.school_class.name}\n{sched.lesson.name}"
                else:
                    cell.value = ""

                cell.font = b_titr
                cell.alignment = rotate_up_wrap

                left_b = thick_side if col_index in day_start_cols else thin_side
                right_b = thick_side if col_index == total_cols else thin_side
                cell.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thick_side)

                col_index += 1

        # خط پررنگ پایین ستون اول (نام دبیر) هم
        cell_teacher.border = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)

        ws.row_dimensions[row_index].height = 60
        row_index += 1

    # خط پررنگ روی هدرها
    for r in [1, 2]:
        col_index = 2
        for day in days:
            for period in day.periods.all():
                cell = ws.cell(row=r, column=col_index)
                left_b = thick_side if col_index in day_start_cols else thin_side
                right_b = thick_side if col_index == total_cols else thin_side
                cell.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thin_side)
                col_index += 1

    # ---------- Column widths ----------
    ws.column_dimensions["A"].width = 16
    max_col = ws.max_column
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5.2

    ws.freeze_panes = "B3"

    # ---------- Response ----------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=barname_dabiran.xlsx"
    wb.save(response)
    return response


def export_schedule_excel_per_class_a5(request):
    """
    هر کلاس یک شیت جداگانه — جدول مثل برنامه کلاسی دیواری:
    ستون اول = روزهای هفته، هدر = زنگ‌ها، تنظیمات چاپ A5
    """
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter

    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------- Styles ----------
    font_title   = Font(name="B Titr", bold=True, size=13)
    font_header  = Font(name="B Titr", bold=True, size=11)
    font_cell    = Font(name="B Titr", size=10)
    font_no_teacher = Font(name="B Titr", size=10, color="FFFFFF", bold=True)

    align_center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right       = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    red_fill    = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    day_fill    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thick = Side(style="medium", color="000000")
    thin  = Side(style="thin",   color="999999")
    border_header = Border(top=thick, bottom=thick, left=thick, right=thick)
    border_day    = Border(top=thin,  bottom=thin,  left=thick, right=thick)
    border_cell   = Border(top=thin,  bottom=thin,  left=thin,  right=thin)
    border_cell_last_col = Border(top=thin, bottom=thin, left=thin, right=thick)

    classes = SchoolClass.objects.filter(school=school).select_related("grade").all()
    days    = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods").order_by("id").all()

    # همه زنگ‌های منحصربه‌فرد (شماره)
    all_periods = []
    seen = set()
    for day in days:
        for p in day.periods.all():
            if p.period_number not in seen:
                seen.add(p.period_number)
                all_periods.append(p)
    all_periods.sort(key=lambda p: p.period_number)
    num_periods = len(all_periods)

    for school_class in classes:
        ws = wb.create_sheet(title=school_class.name[:31])
        ws.sheet_view.rightToLeft = True

        # --------- ردیف ۱: عنوان کلاس ---------
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_periods + 1)
        title_cell = ws.cell(row=1, column=1, value=f"برنامه کلاسی — {school_class.name}")
        title_cell.font    = font_title
        title_cell.alignment = align_center
        title_cell.border  = border_header
        ws.row_dimensions[1].height = 22

        # --------- ردیف ۲: هدر (روزهای هفته / زنگ‌ها) ---------
        corner = ws.cell(row=2, column=1, value="روزهای هفته")
        corner.font      = font_header
        corner.alignment = align_center
        corner.fill      = header_fill
        corner.border    = border_header

        for idx, period in enumerate(all_periods):
            col = idx + 2
            c = ws.cell(row=2, column=col, value=f"زنگ {period.period_number}")
            c.font      = font_header
            c.alignment = align_center
            c.fill      = header_fill
            c.border    = Border(
                top=thick, bottom=thick,
                left=thick if idx == 0 else thin,
                right=thick if idx == num_periods - 1 else thin
            )
        ws.row_dimensions[2].height = 20

        # --------- ردیف‌های بدنه: هر روز یک ردیف ---------
        for row_offset, day in enumerate(days):
            row = row_offset + 3

            # ستون اول: نام روز
            day_cell = ws.cell(row=row, column=1, value=day.name)
            day_cell.font      = font_header
            day_cell.alignment = align_center
            day_cell.fill      = day_fill
            day_cell.border    = border_day

            # دوره‌های روز را به دیکشنری تبدیل کن
            period_map = {p.period_number: p for p in day.periods.all()}

            for idx, period in enumerate(all_periods):
                col = idx + 2
                is_last = (idx == num_periods - 1)
                dp = period_map.get(period.period_number)

                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    top=thin, bottom=thin,
                    left=thin,
                    right=thick if is_last else thin
                )

                if dp is None:
                    cell.value = "—"
                    cell.font      = font_cell
                    cell.alignment = align_center
                    continue

                sched = Schedule.objects.filter(
                    school=school,
                    school_class=school_class,
                    day_period=dp
                ).select_related("lesson", "teacher").first()

                if sched:
                    cell.value     = sched.lesson.name
                    cell.alignment = align_center
                    if sched.teacher is None:
                        cell.fill = red_fill
                        cell.font = font_no_teacher
                    else:
                        cell.font = font_cell
                else:
                    cell.value     = ""
                    cell.font      = font_cell
                    cell.alignment = align_center

            ws.row_dimensions[row].height = 28

        # --------- عرض ستون‌ها ---------
        ws.column_dimensions["A"].width = 13
        period_col_width = max(7, round(110 / num_periods)) if num_periods else 10
        for c in range(2, num_periods + 2):
            ws.column_dimensions[get_column_letter(c)].width = period_col_width

        # --------- تنظیمات چاپ A5 ---------
        ws.page_setup.paperSize  = 11   # A5
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage  = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        ws.print_area   = f"A1:{get_column_letter(num_periods + 1)}{len(list(days)) + 2}"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=barname_per_class_a5.xlsx"
    wb.save(response)
    return response


def export_schedule_excel_per_class(request):
    """هر کلاس یک شیت — روزها در ردیف، زنگ‌ها در ستون، تنظیمات چاپ A5"""
    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    from openpyxl.worksheet.page import PageMargins

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_title      = Font(name="B Titr", bold=True, size=13)
    font_header     = Font(name="B Titr", bold=True, size=11)
    font_cell       = Font(name="B Titr", size=10)
    font_no_teacher = Font(name="B Titr", size=10, color="FFFFFF", bold=True)
    align_center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right     = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    red_fill        = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_fill     = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    day_fill        = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thick           = Side(style="medium", color="000000")
    thin            = Side(style="thin",   color="999999")

    classes = SchoolClass.objects.filter(school=school).select_related("grade").all()
    days    = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods").order_by("id").all()

    all_periods = []
    seen = set()
    for day in days:
        for p in day.periods.all():
            if p.period_number not in seen:
                seen.add(p.period_number)
                all_periods.append(p)
    all_periods.sort(key=lambda p: p.period_number)
    num_periods = len(all_periods)

    for school_class in classes:
        ws = wb.create_sheet(title=school_class.name[:31])
        ws.sheet_view.rightToLeft = True

        # ردیف ۱: عنوان
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_periods + 1)
        tc = ws.cell(row=1, column=1, value=f"برنامه کلاسی — {school_class.name}")
        tc.font = font_title
        tc.alignment = align_center
        tc.border = Border(top=thick, bottom=thick, left=thick, right=thick)
        ws.row_dimensions[1].height = 22

        # ردیف ۲: هدر (روزهای هفته | زنگ‌ها)
        corner = ws.cell(row=2, column=1, value="روزهای هفته")
        corner.font = font_header
        corner.alignment = align_center
        corner.fill = header_fill
        corner.border = Border(top=thick, bottom=thick, left=thick, right=thick)
        for idx, period in enumerate(all_periods):
            col = idx + 2
            c = ws.cell(row=2, column=col, value=f"زنگ {period.period_number}")
            c.font = font_header
            c.alignment = align_center
            c.fill = header_fill
            c.border = Border(
                top=thick, bottom=thick,
                left=thick if idx == 0 else thin,
                right=thick if idx == num_periods - 1 else thin
            )
        ws.row_dimensions[2].height = 20

        # بدنه: هر روز یک ردیف
        for row_offset, day in enumerate(days):
            row = row_offset + 3
            day_cell = ws.cell(row=row, column=1, value=day.name)
            day_cell.font = font_header
            day_cell.alignment = align_center
            day_cell.fill = day_fill
            day_cell.border = Border(top=thin, bottom=thin, left=thick, right=thick)

            period_map = {p.period_number: p for p in day.periods.all()}
            for idx, period in enumerate(all_periods):
                col = idx + 2
                is_last = (idx == num_periods - 1)
                dp = period_map.get(period.period_number)
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thick if is_last else thin)
                if dp is None:
                    cell.value = "—"
                    cell.font = font_cell
                    cell.alignment = align_center
                    continue
                sched = Schedule.objects.filter(
                    school=school, school_class=school_class, day_period=dp
                ).select_related("lesson", "teacher").first()
                if sched:
                    cell.value = sched.lesson.name
                    cell.alignment = align_center
                    if sched.teacher is None:
                        cell.fill = red_fill
                        cell.font = font_no_teacher
                    else:
                        cell.font = font_cell
                else:
                    cell.value = ""
                    cell.font = font_cell
                    cell.alignment = align_center
            ws.row_dimensions[row].height = 28

        ws.column_dimensions["A"].width = 13
        col_w = max(7, round(110 / num_periods)) if num_periods else 10
        for c in range(2, num_periods + 2):
            ws.column_dimensions[get_column_letter(c)].width = col_w

        ws.page_setup.paperSize   = 11
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        ws.print_area = f"A1:{get_column_letter(num_periods + 1)}{len(list(days)) + 2}"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=barname_per_class.xlsx"
    wb.save(response)
    return response


def export_schedule_excel_per_teacher(request):
    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    b_titr = Font(name="B Titr")
    align_right_center = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rotate_up = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=False)
    rotate_up_wrap = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
    thick_side = Side(style="medium", color="000000")
    thin_side = Side(style="thin", color="CCCCCC")

    teachers = Teacher.objects.filter(school=school).all()
    days = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods").all()

    day_start_cols = []
    col_tracker = 2
    for day in days:
        day_start_cols.append(col_tracker)
        col_tracker += len(list(day.periods.all()))
    total_cols = col_tracker - 1

    for teacher in teachers:
        ws = wb.create_sheet(title=teacher.name[:31])
        ws.sheet_view.rightToLeft = True

        # هدر ردیف ۱: روزها
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        cell = ws.cell(row=1, column=1, value="زنگ / روز")
        cell.alignment = align_right_center
        cell.font = b_titr

        col_index = 2
        for day in days:
            periods = list(day.periods.all())
            cell = ws.cell(row=1, column=col_index, value=day.name)
            cell.alignment = align_center
            cell.font = b_titr
            if len(periods) > 1:
                ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + len(periods) - 1)
            col_index += len(periods)

        # هدر ردیف ۲: زنگ‌ها
        col_index = 2
        for day in days:
            for period in day.periods.all():
                cell = ws.cell(row=2, column=col_index, value=f"زنگ {period.period_number}")
                cell.font = b_titr
                cell.alignment = rotate_up
                col_index += 1

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 55

        # بدنه: یک ردیف برای این دبیر
        cell_name = ws.cell(row=3, column=1, value=teacher.name)
        cell_name.alignment = align_right_center
        cell_name.font = b_titr
        cell_name.border = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)

        col_index = 2
        for day in days:
            for period in day.periods.all():
                sched = Schedule.objects.filter(
                    school=school, teacher=teacher, day_period=period
                ).select_related("school_class", "lesson").first()

                cell = ws.cell(row=3, column=col_index)
                if sched:
                    cell.value = f"{sched.school_class.name}\n{sched.lesson.name}"
                else:
                    cell.value = ""

                cell.font = b_titr
                cell.alignment = rotate_up_wrap

                left_b = thick_side if col_index in day_start_cols else thin_side
                right_b = thick_side if col_index == total_cols else thin_side
                cell.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thick_side)
                col_index += 1

        ws.row_dimensions[3].height = 60

        # بوردر هدرها
        for r in [1, 2]:
            col_index = 2
            for day in days:
                for period in day.periods.all():
                    c = ws.cell(row=r, column=col_index)
                    left_b = thick_side if col_index in day_start_cols else thin_side
                    right_b = thick_side if col_index == total_cols else thin_side
                    c.border = Border(left=left_b, right=right_b, top=thin_side, bottom=thick_side)
                    col_index += 1

        ws.column_dimensions["A"].width = 16
        for c in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 5.2
        ws.freeze_panes = "B3"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=barname_per_teacher.xlsx"
    wb.save(response)
    return response


def schedule_table(request):
    school = _require_school(request)
    if school is None:
        return redirect("panel_login")

    classes = SchoolClass.objects.filter(school=school).select_related('grade').all()
    days = SchoolDay.objects.filter(school=school, is_active=True).prefetch_related("periods")

    schedules = Schedule.objects.filter(school=school).select_related(
        'school_class',
        'teacher',
        'lesson',
        'day_period',
        'day_period__day'
    )

    data_table = {}

    for schedule in schedules:
        class_id = schedule.school_class.id
        day_id = schedule.day_period.day.id
        period = schedule.day_period.period_number

        data_table.setdefault(class_id, {})
        data_table[class_id].setdefault(day_id, {})
        data_table[class_id][day_id][period] = schedule

    return render(request, "schedule_table.html", {
        "classes": classes,
        "days": days,
        "data_table": data_table
    })


@require_POST
def build_schedule(request):
    school = _require_school(request)
    if school is None:
        return JsonResponse({"status": "error", "errors": ["No school"]}, status=400)

    errors = run_full_validation()
    if errors:
        return JsonResponse({"status": "error", "errors": errors}, status=400)

    try:
        generate_schedule_with_ortools(school=school)
    except Exception as e:
        return JsonResponse({"status": "error", "errors": [str(e)]}, status=500)

    return JsonResponse({"status": "success", "message": "برنامه با موفقیت تولید شد"})


def schedule_build_view(request):
    school = _require_school(request)
    if school is None:
        return redirect('panel_login')

    if request.method == "POST" and "clear_logs" in request.POST:
        request.session['schedule_logs'] = []
        return redirect('schedule-build')

    if request.method == "POST" and "generate_schedule" in request.POST:
        try:
            logs = generate_schedule_with_ortools(school=school)
            request.session['schedule_logs'] = logs or ["برنامه ساخته شد."]
        except Exception as e:
            request.session['schedule_logs'] = [f"خطا: {e}"]
        return redirect('schedule-build')

    logs = request.session.get('schedule_logs', [])
    return render(request, "schedule_logs.html", {"logs": logs})


def panel_login(request):
    if request.user.is_authenticated:
        return redirect("panel_dashboard")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "✅ با موفقیت وارد شدید.")
            return redirect("panel_dashboard")
        messages.error(request, "❌ نام کاربری یا رمز اشتباه است.")

    return render(request, "auth_login.html")


def panel_register(request):
    if request.user.is_authenticated:
        return redirect("panel_dashboard")

    if request.method == "POST":
        User = get_user_model()

        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        password2 = (request.POST.get("password2") or "").strip()

        school_name = (request.POST.get("school_name") or "").strip()
        school_code = (request.POST.get("school_code") or "").strip()
        education_level = (request.POST.get("education_level") or "").strip()
        manager_mobile = (request.POST.get("manager_mobile") or "").strip()
        manager_full_name = (request.POST.get("manager_full_name") or "").strip()

        if not all([username, password, password2, school_name, school_code, education_level, manager_mobile, manager_full_name]):
            messages.error(request, "❌ لطفاً همه فیلدها را کامل کن.")
            return render(request, "auth_register.html")

        if password != password2:
            messages.error(request, "❌ رمزها یکی نیستند.")
            return render(request, "auth_register.html")

        if User.objects.filter(username=username).exists():
            messages.error(request, "❌ این نام کاربری قبلاً گرفته شده.")
            return render(request, "auth_register.html")

        if School.objects.filter(code=school_code).exists():
            messages.error(request, "❌ این کد آموزشگاه قبلاً ثبت شده.")
            return render(request, "auth_register.html")

        try:
            with transaction.atomic():
                school = School.objects.create(
                    name=school_name,
                    code=school_code,
                    education_level=education_level,
                    manager_full_name=manager_full_name,
                    manager_mobile=manager_mobile,
                )

                user = User.objects.create_user(username=username, password=password)

                if hasattr(user, "school"):
                    user.school = school
                if hasattr(user, "phone"):
                    user.phone = manager_mobile
                if hasattr(user, "full_name"):
                    user.full_name = manager_full_name

                user.save()

            login(request, user)
            messages.success(request, "✅ ثبت‌نام انجام شد.")
            return redirect("panel_dashboard")

        except Exception as e:
            messages.error(request, f"❌ خطا در ثبت‌نام: {e}")

    return render(request, "auth_register.html")


def panel_logout(request):
    logout(request)
    messages.success(request, "✅ خارج شدید.")
    return redirect("panel_login")

from django.http import JsonResponse
from main.models import ScheduleBuildProgress, announce

def company_register(request):
    if request.user.is_authenticated:
        return redirect("company_index")

    if request.method == "POST":
        User = get_user_model()

        full_name = (request.POST.get("full_name") or "").strip()
        mobile = (request.POST.get("mobile") or "").strip()
        email = (request.POST.get("email") or "").strip()
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        password2 = (request.POST.get("password2") or "").strip()

        if not all([full_name, mobile, username, password, password2]):
            messages.error(request, "❌ لطفاً فیلدهای ضروری را کامل کن.")
            return render(request, "site/company_register.html")

        if password != password2:
            messages.error(request, "❌ رمز عبور و تکرار آن یکی نیستند.")
            return render(request, "site/company_register.html")

        if User.objects.filter(username=username).exists():
            messages.error(request, "❌ این نام کاربری قبلاً گرفته شده.")
            return render(request, "site/company_register.html")

        if email and User.objects.filter(email=email).exists():
            messages.error(request, "❌ این ایمیل قبلاً ثبت شده.")
            return render(request, "site/company_register.html")

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    email=email,
                )

                if hasattr(user, "full_name"):
                    user.full_name = full_name

                if hasattr(user, "phone"):
                    user.phone = mobile

                user.save()

            login(request, user)
            messages.success(request, "✅ ثبت‌نام با موفقیت انجام شد.")
            return redirect("company_index")

        except Exception as e:
            messages.error(request, f"❌ خطا در ثبت‌نام: {e}")

    return render(request, "site/company_register.html")


def schedule_progress_api(request):
    school = _require_school(request)
    if not school:
        return JsonResponse({"error": "unauthorized"}, status=403)

    obj, _ = ScheduleBuildProgress.objects.get_or_create(school=school)
    return JsonResponse({
        "percent": obj.percent,
        "status": obj.status,
    })

def company_services(request):
    return render(request, "site/company/services.html")


def company_contact(request):
    return render(request, "site/company/contact.html")


@login_required
def company_dashboard(request):
    return render(request, "site/company/dashboard.html", {
        "orders_count": 0,
        "projects_count": 0,
        "tickets_count": 0,
    })


@login_required
def company_orders(request):
    return render(request, "site/company/orders.html", {
        "orders": [],
    })


@login_required
def company_projects(request):
    return render(request, "site/company/projects.html", {
        "projects": [],
    })



@login_required
def company_tickets(request):
    return render(request, "site/company/tickets.html", {
        "tickets": [],
    })